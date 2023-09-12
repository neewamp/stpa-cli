from UCA import *
import pandas as pd
import openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows
import json 
from typing import Tuple, Optional

def df_to_excel(df, ws, header=True, index=True, startrow=0, startcol=0):
    """Write DataFrame df to openpyxl worksheet ws"""

    rows = dataframe_to_rows(df, header=header, index=index)

    for r_idx, row in enumerate(rows, startrow + 1):
        for c_idx, value in enumerate(row, startcol + 1):
            try: 
                ws.cell(row=r_idx, column=c_idx).value = value
            except:
                pass
#rule_json: str = '{"name":"RL2","control_action":"ABORT_Signal_Weapon_Controller","type":"provided","system":"Weapon_Abort_Controller","contexts":[{"name":"UCA5","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["False","True","True"]},{"name":"UCA6","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["False","True","False"]},{"name":"UCA7","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["False","True","Unknown"]},{"name":"UCA8","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["False","False","True"]},{"name":"UCA9","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["False","False","False"]},{"name":"UCA10","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["False","False","Unknown"]},{"name":"UCA11","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["False","Unknown","True"]},{"name":"UCA12","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["False","Unknown","False"]},{"name":"UCA13","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["False","Unknown","Unknown"]},{"name":"UCA14","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["Unknown","True","True"]},{"name":"UCA15","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["Unknown","True","False"]},{"name":"UCA16","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["Unknown","True","Unknown"]},{"name":"UCA17","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["Unknown","False","True"]},{"name":"UCA18","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["Unknown","False","False"]},{"name":"UCA19","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["Unknown","False","Unknown"]},{"name":"UCA20","hazard_list":["H1","H4"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["Unknown","Unknown","True"]},{"name":"UCA21","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["Unknown","Unknown","False"]},{"name":"UCA22","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["Unknown","Unknown","Unknown"]},{"name":"UCA23","hazard_list":["H1","H4"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["True","False","False"]},{"name":"UCA24","hazard_list":["H1"],"vars":["Safe_Separation","Abort_Command_Recieved","Target_Class_MissMatch"],"values":["True","False","Unknown"]}]}'
try: 
    rules_json: str = input()
    """ rule: Rule = Rule.from_json(rule_json) """
# Person.schema().loads(people_json, many=True)
    rules: List[Rule] = Rule.schema().loads(rules_json, many=True)
    # Just testing pandas and excel 

    """ print(df.to_string()) """
    print("Succesfully parsed rules named: ", ', '.join([rule.name for rule in rules]))
except:
    print("Failed to parse rule")

def rule_to_df_with_context(rule: Rule) -> Tuple[pd.DataFrame, pd.DataFrame]:    
    df = rule_to_df(rule)
    context_df = contexts_to_df(rule.contexts)
    return (df, context_df)
""" df.drop('contexts', axis=1, inplace=True) """
""" print(df.to_string()) """

def rule_to_excel(rule: Rule, sheet_path: str, sheet_name: Optional[str] = None): 
    (df, context_df) = rule_to_df_with_context(rule)

    with pd.ExcelWriter(sheet_path) as writer:  
        df.to_excel(writer, index=False)
        context_df.to_excel(writer, sheet_name='Context-Table', index=False)

def rules_to_excel(rules: List[Rule], sheet_path: str, sheet_name: Optional[str] = None): 
    if (len(rules) <= 0):
        exit(-1)

    with pd.ExcelWriter(sheet_path) as writer:  
        for rule in rules: 
            (df, context_df) = rule_to_df_with_context(rule)
            df_merge = pd.concat([df, context_df])
            print(df_merge.to_string())
            df_merge.to_excel(writer, sheet_name=rule.name, index=False)

rules_to_excel(rules, "output_excel/bleh.xlsx")

""" rule_to_excel(rule, "output_excel/out" + rule.name + ".xlsx") """


""" people_json = '[{"name": "lidatong"}]'
Person.schema().loads(people_json, many=True)  # [Person(name='lidatong')]
 """
""" wb = xl.Workbook() """
""" df_to_excel(df, wb.active) """

""" ws1 = wb.create_sheet("Context-Table")  """
""" df_to_excel(context_df, ws1) """
""" print(context_df.to_string()) """

""" wb.save("output_excel/out" + rule.name + ".xlsx") """
